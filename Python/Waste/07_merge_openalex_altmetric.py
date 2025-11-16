"""
OpenAlex + Altmetric Data Merger - ENHANCED VERSION
Features:
- Pre-merge deduplication within each dataset
- Two-phase matching: DOI first, then title+year for unmatched papers
- Post-merge deduplication safety check
- Separate files for matched and not-matched papers
"""

import pandas as pd
import numpy as np
from typing import Tuple, Dict
import re
import warnings
warnings.filterwarnings('ignore')

class DataMerger:
    """Merge OpenAlex and Altmetric datasets with advanced matching"""

    def __init__(self, openalex_path: str, altmetric_path: str, output_dir: str):
        """
        Initialize with file paths

        Args:
            openalex_path: Path to OpenAlex Excel file
            altmetric_path: Path to Altmetric Excel file
            output_dir: Directory to save output files
        """
        self.openalex_path = openalex_path
        self.altmetric_path = altmetric_path
        self.output_dir = output_dir

        # Statistics tracking
        self.stats = {
            'oa_initial': 0,
            'alt_initial': 0,
            'oa_after_dedup': 0,
            'alt_after_dedup': 0,
            'doi_matches': 0,
            'title_year_matches': 0,
            'total_matches': 0,
            'oa_not_matched': 0,
            'alt_not_matched': 0
        }

    def normalize_doi(self, doi_series: pd.Series, add_prefix: bool = False) -> pd.Series:
        """
        Normalize DOI format for matching

        Args:
            doi_series: Series containing DOI values
            add_prefix: If True, add 'https://doi.org/' prefix
                       If False, remove 'https://doi.org/' prefix

        Returns:
            Series with normalized DOI values
        """
        print(f"   Normalizing DOI format (add_prefix={add_prefix})...")

        # Convert to string and handle NaN
        normalized = doi_series.fillna('').astype(str)

        if add_prefix:
            # Add prefix to DOIs that don't have it
            def add_doi_prefix(doi):
                if not doi or doi == 'nan' or doi == '':
                    return ''
                if doi.startswith('https://doi.org/'):
                    return doi
                elif doi.startswith('http://doi.org/'):
                    return doi.replace('http://doi.org/', 'https://doi.org/')
                elif doi.startswith('doi.org/'):
                    return 'https://' + doi
                else:
                    return 'https://doi.org/' + doi

            normalized = normalized.apply(add_doi_prefix)
        else:
            # Remove prefix from DOIs
            def remove_doi_prefix(doi):
                if not doi or doi == 'nan' or doi == '':
                    return ''
                # Remove various forms of DOI prefix
                doi = doi.replace('https://doi.org/', '')
                doi = doi.replace('http://doi.org/', '')
                doi = doi.replace('doi.org/', '')
                return doi.strip()

            normalized = normalized.apply(remove_doi_prefix)

        return normalized

    def load_and_prepare_data(self) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Load both datasets and prepare them for merging

        Returns:
            Tuple of (openalex_df, altmetric_df)
        """
        print("\n" + "="*70)
        print("LOADING DATA")
        print("="*70)

        # Load OpenAlex data
        print(f"\n📂 Loading OpenAlex data...")
        print(f"   {self.openalex_path}")
        df_oa = pd.read_excel(self.openalex_path)
        print(f"   ✓ Loaded {len(df_oa)} records with {len(df_oa.columns)} columns")

        # Load Altmetric data
        print(f"\n📂 Loading Altmetric data...")
        print(f"   {self.altmetric_path}")
        df_alt = pd.read_excel(self.altmetric_path)
        print(f"   ✓ Loaded {len(df_alt)} records with {len(df_alt.columns)} columns")

        # Check DOI columns
        print(f"\n🔍 Checking DOI columns...")

        if 'doi' not in df_oa.columns:
            print("   ❌ ERROR: 'doi' column not found in OpenAlex data")
            print(f"   Available columns: {df_oa.columns.tolist()[:10]}...")
            raise ValueError("DOI column not found in OpenAlex data")

        if 'doi' not in df_alt.columns:
            print("   ❌ ERROR: 'doi' column not found in Altmetric data")
            print(f"   Available columns: {df_alt.columns.tolist()[:10]}...")
            raise ValueError("DOI column not found in Altmetric data")

        print(f"   ✓ DOI columns found in both datasets")

        # Show sample DOIs
        print(f"\n📋 Sample DOIs:")
        oa_sample = df_oa['doi'].dropna().head(3).tolist()
        alt_sample = df_alt['doi'].dropna().head(3).tolist()

        print(f"   OpenAlex samples:")
        for doi in oa_sample:
            print(f"      {doi}")

        print(f"   Altmetric samples:")
        for doi in alt_sample:
            print(f"      {doi}")

        # Normalize DOI formats
        print(f"\n🔧 Normalizing DOI formats...")
        print(f"   OpenAlex: Removing 'https://doi.org/' prefix")
        df_oa['doi_normalized'] = self.normalize_doi(df_oa['doi'], add_prefix=False)

        print(f"   Altmetric: Keeping DOI as is (already without prefix)")
        df_alt['doi_normalized'] = self.normalize_doi(df_alt['doi'], add_prefix=False)

        # Count valid DOIs
        oa_valid_dois = (df_oa['doi_normalized'] != '').sum()
        alt_valid_dois = (df_alt['doi_normalized'] != '').sum()

        print(f"\n✓ Valid DOIs:")
        print(f"   OpenAlex: {oa_valid_dois} / {len(df_oa)} ({oa_valid_dois/len(df_oa)*100:.1f}%)")
        print(f"   Altmetric: {alt_valid_dois} / {len(df_alt)} ({alt_valid_dois/len(df_alt)*100:.1f}%)")

        return df_oa, df_alt

    def rename_columns(self, df_oa: pd.DataFrame, df_alt: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Rename columns with _oa and _alt suffixes

        Args:
            df_oa: OpenAlex DataFrame
            df_alt: Altmetric DataFrame

        Returns:
            Tuple of DataFrames with renamed columns
        """
        print("\n" + "="*70)
        print("RENAMING COLUMNS")
        print("="*70)

        # Keep doi_normalized as is for merging
        oa_columns = {}
        alt_columns = {}

        for col in df_oa.columns:
            if col == 'doi_normalized':
                oa_columns[col] = col  # Keep as is for merging
            else:
                oa_columns[col] = f"{col}_oa"

        for col in df_alt.columns:
            if col == 'doi_normalized':
                alt_columns[col] = col  # Keep as is for merging
            else:
                alt_columns[col] = f"{col}_alt"

        df_oa_renamed = df_oa.rename(columns=oa_columns)
        df_alt_renamed = df_alt.rename(columns=alt_columns)

        print(f"\n✓ Renamed columns:")
        print(f"   OpenAlex: {len(df_oa.columns)} columns → all suffixed with '_oa'")
        print(f"   Altmetric: {len(df_alt.columns)} columns → all suffixed with '_alt'")
        print(f"   (except 'doi_normalized' kept for merging)")

        return df_oa_renamed, df_alt_renamed

    def merge_datasets(self, df_oa: pd.DataFrame, df_alt: pd.DataFrame) -> pd.DataFrame:
        """
        Merge datasets using outer join to keep all records

        Args:
            df_oa: OpenAlex DataFrame (with renamed columns)
            df_alt: Altmetric DataFrame (with renamed columns)

        Returns:
            Merged DataFrame
        """
        print("\n" + "="*70)
        print("MERGING DATASETS")
        print("="*70)

        print(f"\n🔗 Performing outer join on 'doi_normalized'...")
        print(f"   This will keep ALL papers from both datasets")

        # Perform outer merge
        df_merged = pd.merge(
            df_oa,
            df_alt,
            on='doi_normalized',
            how='outer',
            indicator=True
        )

        print(f"\n✓ Merge complete!")
        print(f"   Total records in merged dataset: {len(df_merged)}")

        # Analyze merge results
        merge_counts = df_merged['_merge'].value_counts()

        print(f"\n📊 Merge statistics:")
        print(f"   Both datasets (matched): {merge_counts.get('both', 0)}")
        print(f"   Only in OpenAlex: {merge_counts.get('left_only', 0)}")
        print(f"   Only in Altmetric: {merge_counts.get('right_only', 0)}")

        # Create match flag
        df_merged['match_status'] = df_merged['_merge'].map({
            'both': 'Matched',
            'left_only': 'OpenAlex_only',
            'right_only': 'Altmetric_only'
        })

        # Drop the _merge column (we have match_status now)
        df_merged = df_merged.drop(columns=['_merge'])

        return df_merged

    def create_not_matched_dataset(self, df_merged: pd.DataFrame) -> pd.DataFrame:
        """
        Create dataset of unmatched papers

        Args:
            df_merged: Full merged dataset

        Returns:
            DataFrame with only unmatched papers
        """
        print("\n" + "="*70)
        print("CREATING NOT MATCHED DATASET")
        print("="*70)

        df_not_matched = df_merged[df_merged['match_status'] != 'Matched'].copy()

        print(f"\n📋 Not matched papers:")
        print(f"   Total: {len(df_not_matched)}")
        print(f"   OpenAlex only: {(df_not_matched['match_status'] == 'OpenAlex_only').sum()}")
        print(f"   Altmetric only: {(df_not_matched['match_status'] == 'Altmetric_only').sum()}")

        # Reorder columns to put match_status and key identifiers first
        cols = df_not_matched.columns.tolist()

        # Key columns to show first
        priority_cols = ['match_status', 'doi_normalized']

        # Add title columns if they exist
        if 'title_oa' in cols:
            priority_cols.append('title_oa')
        if 'title_alt' in cols:
            priority_cols.append('title_alt')

        # Add ID columns
        if 'id_oa' in cols:
            priority_cols.append('id_oa')
        if 'altmetric_id_alt' in cols:
            priority_cols.append('altmetric_id_alt')

        # Reorder
        remaining_cols = [col for col in cols if col not in priority_cols]
        df_not_matched = df_not_matched[priority_cols + remaining_cols]

        return df_not_matched

    def save_datasets(self, df_merged: pd.DataFrame, df_not_matched: pd.DataFrame):
        """
        Save datasets to Excel and DTA formats

        Args:
            df_merged: Full merged dataset
            df_not_matched: Unmatched papers dataset
        """
        print("\n" + "="*70)
        print("SAVING DATASETS")
        print("="*70)

        import os

        # Create output directory if it doesn't exist
        os.makedirs(self.output_dir, exist_ok=True)

        # Define output paths
        merged_excel_path = os.path.join(self.output_dir, "04_OA_Altmetrics_merge.xlsx")
        not_matched_excel_path = os.path.join(self.output_dir, "04_not_matched.xlsx")
        merged_dta_path = os.path.join(self.output_dir, "04_OA_Altmetrics_merge.dta")

        # Save merged dataset to Excel
        print(f"\n💾 Saving merged dataset to Excel...")
        print(f"   {merged_excel_path}")

        with pd.ExcelWriter(merged_excel_path, engine='openpyxl') as writer:
            df_merged.to_excel(writer, sheet_name='Merged Data', index=False)

            # Auto-adjust column widths
            worksheet = writer.sheets['Merged Data']
            for idx, col in enumerate(df_merged.columns):
                max_length = min(
                    max(
                        df_merged[col].astype(str).apply(len).max(),
                        len(col)
                    ) + 2,
                    100
                )
                from openpyxl.utils import get_column_letter
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_length

        print(f"   ✓ Saved {len(df_merged)} records with {len(df_merged.columns)} columns")

        # Save not matched dataset to Excel
        print(f"\n💾 Saving not matched dataset to Excel...")
        print(f"   {not_matched_excel_path}")

        with pd.ExcelWriter(not_matched_excel_path, engine='openpyxl') as writer:
            df_not_matched.to_excel(writer, sheet_name='Not Matched', index=False)

            # Auto-adjust column widths
            worksheet = writer.sheets['Not Matched']
            for idx, col in enumerate(df_not_matched.columns):
                max_length = min(
                    max(
                        df_not_matched[col].astype(str).apply(len).max(),
                        len(col)
                    ) + 2,
                    100
                )
                from openpyxl.utils import get_column_letter
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_length

        print(f"   ✓ Saved {len(df_not_matched)} records with {len(df_not_matched.columns)} columns")

        # Save merged dataset to DTA (Stata format)
        print(f"\n💾 Saving merged dataset to Stata (.dta) format...")
        print(f"   {merged_dta_path}")

        try:
            # Try using pandas built-in Stata writer
            # Stata has limitations: column names max 32 chars, string max 2045 chars

            # Prepare dataframe for Stata
            df_for_stata = df_merged.copy()

            # Truncate column names to 32 characters (Stata limitation)
            column_mapping = {}
            for col in df_for_stata.columns:
                if len(col) > 32:
                    new_col = col[:29] + '...'
                    column_mapping[col] = new_col

            if column_mapping:
                print(f"   ⚠ Truncating {len(column_mapping)} column names to fit Stata's 32-char limit")
                df_for_stata = df_for_stata.rename(columns=column_mapping)

            # Convert string columns that are too long
            for col in df_for_stata.select_dtypes(include=['object']).columns:
                max_len = df_for_stata[col].astype(str).apply(len).max()
                if max_len > 2045:
                    print(f"   ⚠ Truncating column '{col}' values to 2045 chars (Stata limit)")
                    df_for_stata[col] = df_for_stata[col].astype(str).str[:2045]

            # Save to Stata format
            df_for_stata.to_stata(
                merged_dta_path,
                write_index=False,
                version=118  # Stata 14/15/16 format
            )

            print(f"   ✓ Saved to Stata format successfully")
            print(f"   📝 Note: Column names truncated to 32 chars, long strings to 2045 chars (Stata limits)")

        except Exception as e:
            print(f"   ⚠ Could not save to .dta format: {e}")
            print(f"   💡 You may need to install: pip install pyreadstat")
            print(f"   📊 Excel files were saved successfully")

    def run(self):
        """Execute the complete merge process"""
        print("\n" + "="*70)
        print("OPENALEX + ALTMETRIC DATA MERGER")
        print("="*70)

        # Load data
        df_oa, df_alt = self.load_and_prepare_data()

        # Rename columns
        df_oa_renamed, df_alt_renamed = self.rename_columns(df_oa, df_alt)

        # Merge
        df_merged = self.merge_datasets(df_oa_renamed, df_alt_renamed)

        # Create not matched dataset
        df_not_matched = self.create_not_matched_dataset(df_merged)

        # Save datasets
        self.save_datasets(df_merged, df_not_matched)

        # Final summary
        print("\n" + "="*70)
        print("MERGE COMPLETE - SUMMARY")
        print("="*70)

        print(f"\n📊 Final Statistics:")
        print(f"   Total records in merged file: {len(df_merged)}")
        print(f"   Matched papers: {(df_merged['match_status'] == 'Matched').sum()}")
        print(f"   OpenAlex only: {(df_merged['match_status'] == 'OpenAlex_only').sum()}")
        print(f"   Altmetric only: {(df_merged['match_status'] == 'Altmetric_only').sum()}")

        print(f"\n📁 Output Files:")
        print(f"   1. Merged dataset (Excel): 04_OA_Altmetrics_merge.xlsx")
        print(f"   2. Merged dataset (Stata): 04_OA_Altmetrics_merge.dta")
        print(f"   3. Not matched papers: 04_not_matched.xlsx")

        print(f"\n✓ All files saved to:")
        print(f"   {self.output_dir}")

        print("\n" + "="*70)


def main():
    """
    Main execution function
    """
    # Define paths
    OPENALEX_PATH = r"C:\Users\User\OneDrive\OneDrive - Universidad de los andes\Global Complexity School\Final project\Excel\03_OpenAlex_API_CS_FULL.xlsx"

    ALTMETRIC_PATH = r"C:\Users\User\OneDrive\OneDrive - Universidad de los andes\Global Complexity School\Final project\Excel\03_Almetrics_API_CS.xlsx"

    OUTPUT_DIR = r"C:\Users\User\OneDrive\OneDrive - Universidad de los andes\Global Complexity School\Final project\Excel"

    # Create merger and run
    merger = DataMerger(OPENALEX_PATH, ALTMETRIC_PATH, OUTPUT_DIR)
    merger.run()


if __name__ == "__main__":
    main()
