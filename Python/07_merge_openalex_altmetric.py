"""
OpenAlex + Altmetric Data Merger - ENHANCED VERSION
Features:
- Pre-merge deduplication within each dataset
- Two-phase matching: DOI first, then title+year for unmatched papers
- Post-merge deduplication safety check
- Separate files for matched and not-matched papers
"""

import pandas as pd
import numpy as np
from typing import Tuple
import re
import warnings
warnings.filterwarnings('ignore')

class DataMerger:
    """Merge OpenAlex and Altmetric datasets with advanced matching"""

    def __init__(self, openalex_path: str, altmetric_path: str, output_dir: str):
        """
        Initialize with file paths

        Args:
            openalex_path: Path to OpenAlex Excel file
            altmetric_path: Path to Altmetric Excel file
            output_dir: Directory to save output files
        """
        self.openalex_path = openalex_path
        self.altmetric_path = altmetric_path
        self.output_dir = output_dir

        # Statistics tracking
        self.stats = {
            'oa_initial': 0,
            'alt_initial': 0,
            'oa_after_dedup': 0,
            'alt_after_dedup': 0,
            'doi_matches': 0,
            'title_year_matches': 0,
            'total_matches': 0,
            'oa_not_matched': 0,
            'alt_not_matched': 0
        }

    def normalize_doi(self, doi_series: pd.Series) -> pd.Series:
        """
        Normalize DOI format by removing URL prefix

        Args:
            doi_series: Series containing DOI values

        Returns:
            Series with normalized DOI values (lowercase, no prefix)
        """
        def clean_doi(doi):
            if pd.isna(doi) or doi == '' or str(doi) == 'nan':
                return ''

            doi_str = str(doi).strip()

            # Remove various forms of DOI prefix
            doi_str = doi_str.replace('https://doi.org/', '')
            doi_str = doi_str.replace('http://doi.org/', '')
            doi_str = doi_str.replace('doi.org/', '')
            doi_str = doi_str.replace('DOI:', '')
            doi_str = doi_str.replace('doi:', '')

            # Lowercase and strip
            doi_str = doi_str.lower().strip()

            return doi_str if doi_str else ''

        return doi_series.apply(clean_doi)

    def normalize_title(self, title_series: pd.Series) -> pd.Series:
        """
        Aggressively normalize titles for matching

        Strategy: Exact match after heavy normalization to avoid false positives

        Steps:
        - Convert to lowercase
        - Remove all punctuation and special characters
        - Remove extra whitespace
        - Keep only alphanumeric and spaces

        Args:
            title_series: Series containing title values

        Returns:
            Series with normalized title values
        """
        def clean_title(title):
            if pd.isna(title) or title == '' or str(title) == 'nan':
                return ''

            title_str = str(title).strip()

            # Convert to lowercase
            title_str = title_str.lower()

            # Remove common HTML entities
            title_str = title_str.replace('&amp;', 'and')
            title_str = title_str.replace('&nbsp;', ' ')

            # Remove all punctuation and special characters
            # Keep only letters, numbers, and spaces
            title_str = re.sub(r'[^a-z0-9\s]', ' ', title_str)

            # Remove extra whitespace
            title_str = re.sub(r'\s+', ' ', title_str)

            # Strip leading/trailing spaces
            title_str = title_str.strip()

            return title_str if title_str else ''

        return title_series.apply(clean_title)

    def extract_year(self, date_series: pd.Series) -> pd.Series:
        """
        Extract year from date field

        Args:
            date_series: Series containing date values (various formats)

        Returns:
            Series with extracted years as integers
        """
        def get_year(date_val):
            if pd.isna(date_val) or date_val == '' or str(date_val) == 'nan':
                return np.nan

            date_str = str(date_val)

            # Try to extract 4-digit year
            year_match = re.search(r'\b(19|20)\d{2}\b', date_str)
            if year_match:
                return int(year_match.group())

            return np.nan

        return date_series.apply(get_year)

    def deduplicate_dataset(self, df: pd.DataFrame, dataset_name: str,
                           doi_col: str, title_col: str, year_col: str) -> pd.DataFrame:
        """
        Deduplicate a dataset by DOI first, then by title+year

        Strategy: Keep first occurrence in both cases

        Args:
            df: DataFrame to deduplicate
            dataset_name: Name for logging (e.g., "OpenAlex")
            doi_col: Name of DOI column
            title_col: Name of title column
            year_col: Name of year column

        Returns:
            Deduplicated DataFrame
        """
        print(f"\n{'─'*70}")
        print(f"DEDUPLICATING {dataset_name}")
        print(f"{'─'*70}")

        initial_count = len(df)
        print(f"   Initial records: {initial_count}")

        # Create normalized columns for deduplication
        df['_doi_norm'] = self.normalize_doi(df[doi_col])
        df['_title_norm'] = self.normalize_title(df[title_col])

        # Step 1: Remove duplicates by DOI (where DOI exists)
        # Keep first occurrence
        df_with_doi = df[df['_doi_norm'] != ''].copy()
        df_without_doi = df[df['_doi_norm'] == ''].copy()

        before_doi_dedup = len(df_with_doi)
        df_with_doi = df_with_doi.drop_duplicates(subset=['_doi_norm'], keep='first')
        after_doi_dedup = len(df_with_doi)
        doi_duplicates_removed = before_doi_dedup - after_doi_dedup

        print(f"   DOI-based deduplication:")
        print(f"      Records with DOI: {before_doi_dedup}")
        print(f"      Duplicates removed: {doi_duplicates_removed}")
        print(f"      Records after: {after_doi_dedup}")

        # Step 2: Remove duplicates by title+year (for remaining papers)
        # Combine back
        df = pd.concat([df_with_doi, df_without_doi], ignore_index=True)

        # Remove duplicates by title+year (where both exist)
        df_with_title_year = df[(df['_title_norm'] != '') & (df[year_col].notna())].copy()
        df_without_title_year = df[~((df['_title_norm'] != '') & (df[year_col].notna()))].copy()

        before_title_dedup = len(df_with_title_year)
        df_with_title_year = df_with_title_year.drop_duplicates(
            subset=['_title_norm', year_col], keep='first'
        )
        after_title_dedup = len(df_with_title_year)
        title_duplicates_removed = before_title_dedup - after_title_dedup

        print(f"   Title+Year deduplication:")
        print(f"      Records with title+year: {before_title_dedup}")
        print(f"      Duplicates removed: {title_duplicates_removed}")
        print(f"      Records after: {after_title_dedup}")

        # Combine back
        df = pd.concat([df_with_title_year, df_without_title_year], ignore_index=True)

        final_count = len(df)
        total_removed = initial_count - final_count

        print(f"   ✓ Final records: {final_count}")
        print(f"   ✓ Total duplicates removed: {total_removed} ({total_removed/initial_count*100:.1f}%)")

        return df

    def load_and_prepare_data(self) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Load both datasets and prepare them for merging

        Returns:
            Tuple of (openalex_df, altmetric_df) - both deduplicated
        """
        print("\n" + "="*70)
        print("STEP 1: LOADING DATA")
        print("="*70)

        # Load OpenAlex data
        print(f"\n📂 Loading OpenAlex data...")
        print(f"   {self.openalex_path}")
        df_oa = pd.read_excel(self.openalex_path)
        self.stats['oa_initial'] = len(df_oa)
        print(f"   ✓ Loaded {len(df_oa)} records with {len(df_oa.columns)} columns")

        # Load Altmetric data
        print(f"\n📂 Loading Altmetric data...")
        print(f"   {self.altmetric_path}")
        df_alt = pd.read_excel(self.altmetric_path)
        self.stats['alt_initial'] = len(df_alt)
        print(f"   ✓ Loaded {len(df_alt)} records with {len(df_alt.columns)} columns")

        # Validate required columns
        print(f"\n🔍 Validating required columns...")

        required_oa_cols = ['doi', 'title', 'publication_year']
        required_alt_cols = ['doi', 'title', 'publication_date']

        missing_oa = [col for col in required_oa_cols if col not in df_oa.columns]
        missing_alt = [col for col in required_alt_cols if col not in df_alt.columns]

        if missing_oa:
            raise ValueError(f"Missing columns in OpenAlex: {missing_oa}")
        if missing_alt:
            raise ValueError(f"Missing columns in Altmetric: {missing_alt}")

        print(f"   ✓ All required columns present")

        # Extract year from Altmetric publication_date
        print(f"\n📅 Extracting year from Altmetric publication_date...")
        df_alt['publication_year'] = self.extract_year(df_alt['publication_date'])
        print(f"   ✓ Extracted years for {df_alt['publication_year'].notna().sum()} / {len(df_alt)} records")

        # Show sample data
        print(f"\n📋 Sample data:")
        print(f"   OpenAlex DOI: {df_oa['doi'].dropna().head(2).tolist()}")
        print(f"   Altmetric DOI: {df_alt['doi'].dropna().head(2).tolist()}")

        print("\n" + "="*70)
        print("STEP 2: PRE-MERGE DEDUPLICATION")
        print("="*70)

        # Deduplicate OpenAlex
        df_oa = self.deduplicate_dataset(
            df_oa, "OPENALEX",
            doi_col='doi',
            title_col='title',
            year_col='publication_year'
        )
        self.stats['oa_after_dedup'] = len(df_oa)

        # Deduplicate Altmetric
        df_alt = self.deduplicate_dataset(
            df_alt, "ALTMETRIC",
            doi_col='doi',
            title_col='title',
            year_col='publication_year'
        )
        self.stats['alt_after_dedup'] = len(df_alt)

        print(f"\n{'='*70}")
        print(f"DEDUPLICATION SUMMARY")
        print(f"{'='*70}")
        print(f"   OpenAlex: {self.stats['oa_initial']} → {self.stats['oa_after_dedup']} "
              f"(-{self.stats['oa_initial'] - self.stats['oa_after_dedup']})")
        print(f"   Altmetric: {self.stats['alt_initial']} → {self.stats['alt_after_dedup']} "
              f"(-{self.stats['alt_initial'] - self.stats['alt_after_dedup']})")

        return df_oa, df_alt

    def rename_columns(self, df_oa: pd.DataFrame, df_alt: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """
        Rename columns with _oa and _alt suffixes
        Keep normalized columns for matching
        """
        print("\n" + "="*70)
        print("STEP 3: RENAMING COLUMNS")
        print("="*70)

        # Columns to keep without suffix (for matching)
        merge_keys = ['_doi_norm', '_title_norm', 'publication_year']

        oa_columns = {}
        alt_columns = {}

        for col in df_oa.columns:
            if col in merge_keys:
                oa_columns[col] = col
            else:
                oa_columns[col] = f"{col}_oa"

        for col in df_alt.columns:
            if col in merge_keys:
                alt_columns[col] = col
            else:
                alt_columns[col] = f"{col}_alt"

        df_oa_renamed = df_oa.rename(columns=oa_columns)
        df_alt_renamed = df_alt.rename(columns=alt_columns)

        print(f"   ✓ OpenAlex: {len(df_oa.columns)} columns → suffixed with '_oa'")
        print(f"   ✓ Altmetric: {len(df_alt.columns)} columns → suffixed with '_alt'")
        print(f"   ✓ Merge keys preserved: {merge_keys}")

        return df_oa_renamed, df_alt_renamed

    def merge_by_doi(self, df_oa: pd.DataFrame, df_alt: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """
        Phase 1: Merge by DOI

        Returns:
            Tuple of (matched_df, unmatched_oa_df, unmatched_alt_df)
        """
        print("\n" + "="*70)
        print("STEP 4: PHASE 1 - MERGING BY DOI")
        print("="*70)

        # Only merge papers with valid DOIs
        df_oa_with_doi = df_oa[df_oa['_doi_norm'] != ''].copy()
        df_alt_with_doi = df_alt[df_alt['_doi_norm'] != ''].copy()

        df_oa_no_doi = df_oa[df_oa['_doi_norm'] == ''].copy()
        df_alt_no_doi = df_alt[df_alt['_doi_norm'] == ''].copy()

        print(f"   OpenAlex papers with DOI: {len(df_oa_with_doi)}")
        print(f"   Altmetric papers with DOI: {len(df_alt_with_doi)}")

        # Merge by DOI
        df_doi_matched = pd.merge(
            df_oa_with_doi,
            df_alt_with_doi,
            on='_doi_norm',
            how='inner',
            suffixes=('', '_DROP')
        )

        # Drop duplicate year column from merge
        drop_cols = [col for col in df_doi_matched.columns if col.endswith('_DROP')]
        df_doi_matched = df_doi_matched.drop(columns=drop_cols)

        df_doi_matched['match_type'] = 'DOI_match'

        self.stats['doi_matches'] = len(df_doi_matched)

        print(f"   ✓ DOI matches found: {len(df_doi_matched)}")

        # Find unmatched papers
        matched_oa_dois = set(df_doi_matched['_doi_norm'].unique())
        matched_alt_dois = set(df_doi_matched['_doi_norm'].unique())

        df_oa_unmatched = pd.concat([
            df_oa_with_doi[~df_oa_with_doi['_doi_norm'].isin(matched_oa_dois)],
            df_oa_no_doi
        ], ignore_index=True)

        df_alt_unmatched = pd.concat([
            df_alt_with_doi[~df_alt_with_doi['_doi_norm'].isin(matched_alt_dois)],
            df_alt_no_doi
        ], ignore_index=True)

        print(f"   Unmatched OpenAlex papers: {len(df_oa_unmatched)}")
        print(f"   Unmatched Altmetric papers: {len(df_alt_unmatched)}")

        return df_doi_matched, df_oa_unmatched, df_alt_unmatched

    def merge_by_title_year(self, df_oa: pd.DataFrame, df_alt: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """
        Phase 2: Merge by normalized title + year

        Args:
            df_oa: Unmatched OpenAlex papers from Phase 1
            df_alt: Unmatched Altmetric papers from Phase 1

        Returns:
            Tuple of (matched_df, still_unmatched_oa_df, still_unmatched_alt_df)
        """
        print("\n" + "="*70)
        print("STEP 5: PHASE 2 - MERGING BY TITLE + YEAR")
        print("="*70)

        # Only merge papers with valid title and year
        df_oa_valid = df_oa[
            (df_oa['_title_norm'] != '') &
            (df_oa['publication_year'].notna())
        ].copy()

        df_alt_valid = df_alt[
            (df_alt['_title_norm'] != '') &
            (df_alt['publication_year'].notna())
        ].copy()

        df_oa_invalid = df_oa[
            ~((df_oa['_title_norm'] != '') & (df_oa['publication_year'].notna()))
        ].copy()

        df_alt_invalid = df_alt[
            ~((df_alt['_title_norm'] != '') & (df_alt['publication_year'].notna()))
        ].copy()

        print(f"   OpenAlex papers with title+year: {len(df_oa_valid)}")
        print(f"   Altmetric papers with title+year: {len(df_alt_valid)}")

        # Merge by title + year
        df_title_matched = pd.merge(
            df_oa_valid,
            df_alt_valid,
            on=['_title_norm', 'publication_year'],
            how='inner'
        )

        df_title_matched['match_type'] = 'Title_Year_match'

        self.stats['title_year_matches'] = len(df_title_matched)

        print(f"   ✓ Title+Year matches found: {len(df_title_matched)}")

        # Find still unmatched papers
        matched_oa_keys = set(
            df_title_matched['_title_norm'] + '|' + df_title_matched['publication_year'].astype(str)
        )
        matched_alt_keys = set(
            df_title_matched['_title_norm'] + '|' + df_title_matched['publication_year'].astype(str)
        )

        df_oa_valid['_match_key'] = df_oa_valid['_title_norm'] + '|' + df_oa_valid['publication_year'].astype(str)
        df_alt_valid['_match_key'] = df_alt_valid['_title_norm'] + '|' + df_alt_valid['publication_year'].astype(str)

        df_oa_still_unmatched = pd.concat([
            df_oa_valid[~df_oa_valid['_match_key'].isin(matched_oa_keys)].drop(columns=['_match_key']),
            df_oa_invalid
        ], ignore_index=True)

        df_alt_still_unmatched = pd.concat([
            df_alt_valid[~df_alt_valid['_match_key'].isin(matched_alt_keys)].drop(columns=['_match_key']),
            df_alt_invalid
        ], ignore_index=True)

        print(f"   Still unmatched OpenAlex papers: {len(df_oa_still_unmatched)}")
        print(f"   Still unmatched Altmetric papers: {len(df_alt_still_unmatched)}")

        return df_title_matched, df_oa_still_unmatched, df_alt_still_unmatched

    def post_merge_deduplication(self, df: pd.DataFrame, label: str) -> pd.DataFrame:
        """
        Safety check: Remove any remaining duplicates after merging
        Priority: DOI match > Title match, Earlier rows > Later rows

        Args:
            df: DataFrame to deduplicate
            label: Label for logging

        Returns:
            Deduplicated DataFrame
        """
        print(f"\n{'─'*70}")
        print(f"POST-MERGE DEDUPLICATION: {label}")
        print(f"{'─'*70}")

        initial_count = len(df)
        print(f"   Initial records: {initial_count}")

        # Sort by match_type priority (DOI_match first) if applicable
        if 'match_type' in df.columns:
            df = df.sort_values('match_type', ascending=True)  # DOI_match < Title_Year_match

        # Step 1: Remove duplicates by DOI (where DOI exists)
        df_with_doi = df[df['_doi_norm'] != ''].copy()
        df_without_doi = df[df['_doi_norm'] == ''].copy()

        before_doi = len(df_with_doi)
        df_with_doi = df_with_doi.drop_duplicates(subset=['_doi_norm'], keep='first')
        after_doi = len(df_with_doi)

        print(f"   DOI deduplication: {before_doi} → {after_doi} (-{before_doi - after_doi})")

        # Step 2: Remove duplicates by title+year
        df = pd.concat([df_with_doi, df_without_doi], ignore_index=True)

        df_with_title_year = df[(df['_title_norm'] != '') & (df['publication_year'].notna())].copy()
        df_without_title_year = df[~((df['_title_norm'] != '') & (df['publication_year'].notna()))].copy()

        before_title = len(df_with_title_year)
        df_with_title_year = df_with_title_year.drop_duplicates(
            subset=['_title_norm', 'publication_year'], keep='first'
        )
        after_title = len(df_with_title_year)

        print(f"   Title+Year deduplication: {before_title} → {after_title} (-{before_title - after_title})")

        df = pd.concat([df_with_title_year, df_without_title_year], ignore_index=True)

        final_count = len(df)
        print(f"   ✓ Final records: {final_count} (removed {initial_count - final_count})")

        return df

    def create_not_matched_dataset(self, df_oa_unmatched: pd.DataFrame,
                                   df_alt_unmatched: pd.DataFrame) -> pd.DataFrame:
        """
        Create dataset of unmatched papers from both sources

        Args:
            df_oa_unmatched: OpenAlex papers that didn't match
            df_alt_unmatched: Altmetric papers that didn't match

        Returns:
            Combined DataFrame with match_status flag
        """
        print("\n" + "="*70)
        print("STEP 7: CREATING NOT-MATCHED DATASET")
        print("="*70)

        # Add source indicators
        df_oa_unmatched = df_oa_unmatched.copy()
        df_alt_unmatched = df_alt_unmatched.copy()

        df_oa_unmatched['match_status'] = 'OpenAlex_only'
        df_alt_unmatched['match_status'] = 'Altmetric_only'

        # Get all columns from both datasets
        all_cols = list(set(df_oa_unmatched.columns) | set(df_alt_unmatched.columns))

        # Ensure both have same columns (fill missing with NaN)
        for col in all_cols:
            if col not in df_oa_unmatched.columns:
                df_oa_unmatched[col] = np.nan
            if col not in df_alt_unmatched.columns:
                df_alt_unmatched[col] = np.nan

        # Combine
        df_not_matched = pd.concat([df_oa_unmatched, df_alt_unmatched], ignore_index=True)

        # Deduplicate (should be minimal at this point)
        df_not_matched = self.post_merge_deduplication(df_not_matched, "NOT MATCHED")

        self.stats['oa_not_matched'] = (df_not_matched['match_status'] == 'OpenAlex_only').sum()
        self.stats['alt_not_matched'] = (df_not_matched['match_status'] == 'Altmetric_only').sum()

        print(f"\n   ✓ Not matched papers:")
        print(f"      OpenAlex only: {self.stats['oa_not_matched']}")
        print(f"      Altmetric only: {self.stats['alt_not_matched']}")
        print(f"      Total: {len(df_not_matched)}")

        # Reorder columns - put key info first
        priority_cols = ['match_status', '_doi_norm', '_title_norm', 'publication_year']

        if 'title_oa' in df_not_matched.columns:
            priority_cols.append('title_oa')
        if 'title_alt' in df_not_matched.columns:
            priority_cols.append('title_alt')
        if 'doi_oa' in df_not_matched.columns:
            priority_cols.append('doi_oa')
        if 'doi_alt' in df_not_matched.columns:
            priority_cols.append('doi_alt')

        remaining_cols = [col for col in df_not_matched.columns if col not in priority_cols]
        df_not_matched = df_not_matched[[col for col in priority_cols if col in df_not_matched.columns] + remaining_cols]

        return df_not_matched

    def save_datasets(self, df_matched: pd.DataFrame, df_not_matched: pd.DataFrame):
        """
        Save datasets to Excel and DTA formats
        """
        print("\n" + "="*70)
        print("STEP 8: SAVING DATASETS")
        print("="*70)

        import os
        from openpyxl.utils import get_column_letter

        # Create output directory if needed
        os.makedirs(self.output_dir, exist_ok=True)

        # Define paths
        matched_excel_path = os.path.join(self.output_dir, "04_OA_Altmetrics_merge.xlsx")
        not_matched_excel_path = os.path.join(self.output_dir, "04_not_matched.xlsx")
        matched_dta_path = os.path.join(self.output_dir, "04_OA_Altmetrics_merge.dta")

        # Clean up internal columns before saving
        internal_cols = ['_doi_norm', '_title_norm', '_match_key']
        df_matched_clean = df_matched.drop(columns=[col for col in internal_cols if col in df_matched.columns])
        df_not_matched_clean = df_not_matched.drop(columns=[col for col in internal_cols if col in df_not_matched.columns])

        # Save matched dataset to Excel
        print(f"\n💾 Saving MATCHED dataset to Excel...")
        print(f"   {matched_excel_path}")

        with pd.ExcelWriter(matched_excel_path, engine='openpyxl') as writer:
            df_matched_clean.to_excel(writer, sheet_name='Matched Papers', index=False)

            worksheet = writer.sheets['Matched Papers']
            for idx, col in enumerate(df_matched_clean.columns):
                max_length = min(
                    max(
                        df_matched_clean[col].astype(str).apply(len).max(),
                        len(col)
                    ) + 2,
                    100
                )
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_length

        print(f"   ✓ Saved {len(df_matched_clean)} matched papers with {len(df_matched_clean.columns)} columns")

        # Save not-matched dataset to Excel
        print(f"\n💾 Saving NOT-MATCHED dataset to Excel...")
        print(f"   {not_matched_excel_path}")

        with pd.ExcelWriter(not_matched_excel_path, engine='openpyxl') as writer:
            df_not_matched_clean.to_excel(writer, sheet_name='Not Matched Papers', index=False)

            worksheet = writer.sheets['Not Matched Papers']
            for idx, col in enumerate(df_not_matched_clean.columns):
                max_length = min(
                    max(
                        df_not_matched_clean[col].astype(str).apply(len).max(),
                        len(col)
                    ) + 2,
                    100
                )
                worksheet.column_dimensions[get_column_letter(idx + 1)].width = max_length

        print(f"   ✓ Saved {len(df_not_matched_clean)} not-matched papers with {len(df_not_matched_clean.columns)} columns")

        # Save to Stata format
        print(f"\n💾 Saving MATCHED dataset to Stata (.dta)...")
        print(f"   {matched_dta_path}")

        try:
            df_for_stata = df_matched_clean.copy()

            # Truncate column names to 32 characters
            column_mapping = {}
            for col in df_for_stata.columns:
                if len(col) > 32:
                    new_col = col[:29] + '...'
                    counter = 1
                    while new_col in column_mapping.values():
                        new_col = col[:27] + f'{counter:02d}'
                        counter += 1
                    column_mapping[col] = new_col

            if column_mapping:
                print(f"   ⚠ Truncating {len(column_mapping)} column names to 32 chars (Stata limit)")
                df_for_stata = df_for_stata.rename(columns=column_mapping)

            # Truncate long strings
            for col in df_for_stata.select_dtypes(include=['object']).columns:
                df_for_stata[col] = df_for_stata[col].astype(str).str[:2045]

            df_for_stata.to_stata(matched_dta_path, write_index=False, version=118)

            print(f"   ✓ Saved to Stata format")
            print(f"   📝 Note: Column names truncated to 32 chars, strings to 2045 chars")

        except Exception as e:
            print(f"   ⚠ Could not save .dta: {e}")
            print(f"   💡 Install with: pip install pyreadstat")

    def run(self):
        """Execute the complete merge process"""
        print("\n" + "="*70)
        print("OPENALEX + ALTMETRIC MERGER - ENHANCED VERSION")
        print("="*70)
        print("Features:")
        print("  • Pre-merge deduplication")
        print("  • DOI matching (Phase 1)")
        print("  • Title+Year matching (Phase 2)")
        print("  • Post-merge deduplication")
        print("  • Separate matched/not-matched files")
        print("="*70)

        # Step 1-2: Load and deduplicate
        df_oa, df_alt = self.load_and_prepare_data()

        # Step 3: Rename columns
        df_oa, df_alt = self.rename_columns(df_oa, df_alt)

        # Step 4: Phase 1 - DOI matching
        df_doi_matched, df_oa_unmatched, df_alt_unmatched = self.merge_by_doi(df_oa, df_alt)

        # Step 5: Phase 2 - Title+Year matching
        df_title_matched, df_oa_still_unmatched, df_alt_still_unmatched = \
            self.merge_by_title_year(df_oa_unmatched, df_alt_unmatched)

        # Step 6: Combine all matches and deduplicate
        print("\n" + "="*70)
        print("STEP 6: COMBINING MATCHES")
        print("="*70)

        df_all_matched = pd.concat([df_doi_matched, df_title_matched], ignore_index=True)
        print(f"   Total matches before deduplication: {len(df_all_matched)}")
        print(f"      DOI matches: {len(df_doi_matched)}")
        print(f"      Title+Year matches: {len(df_title_matched)}")

        # Post-merge deduplication
        df_all_matched = self.post_merge_deduplication(df_all_matched, "MATCHED PAPERS")

        self.stats['total_matches'] = len(df_all_matched)

        # Step 7: Create not-matched dataset
        df_not_matched = self.create_not_matched_dataset(df_oa_still_unmatched, df_alt_still_unmatched)

        # Step 8: Save
        self.save_datasets(df_all_matched, df_not_matched)

        # Final summary
        self.print_final_summary()

    def print_final_summary(self):
        """Print comprehensive final summary"""
        print("\n" + "="*70)
        print("FINAL SUMMARY")
        print("="*70)

        print(f"\n📊 INPUT DATASETS:")
        print(f"   OpenAlex initial: {self.stats['oa_initial']}")
        print(f"   Altmetric initial: {self.stats['alt_initial']}")

        print(f"\n🔧 AFTER DEDUPLICATION:")
        print(f"   OpenAlex: {self.stats['oa_after_dedup']} "
              f"(-{self.stats['oa_initial'] - self.stats['oa_after_dedup']})")
        print(f"   Altmetric: {self.stats['alt_after_dedup']} "
              f"(-{self.stats['alt_initial'] - self.stats['alt_after_dedup']})")

        print(f"\n✅ MATCHING RESULTS:")
        print(f"   DOI matches: {self.stats['doi_matches']}")
        print(f"   Title+Year matches: {self.stats['title_year_matches']}")
        print(f"   Total matches: {self.stats['total_matches']}")

        match_rate_oa = self.stats['total_matches'] / self.stats['oa_after_dedup'] * 100 if self.stats['oa_after_dedup'] > 0 else 0
        match_rate_alt = self.stats['total_matches'] / self.stats['alt_after_dedup'] * 100 if self.stats['alt_after_dedup'] > 0 else 0

        print(f"   Match rate (OpenAlex): {match_rate_oa:.1f}%")
        print(f"   Match rate (Altmetric): {match_rate_alt:.1f}%")

        print(f"\n❌ NOT MATCHED:")
        print(f"   OpenAlex only: {self.stats['oa_not_matched']}")
        print(f"   Altmetric only: {self.stats['alt_not_matched']}")
        print(f"   Total not matched: {self.stats['oa_not_matched'] + self.stats['alt_not_matched']}")

        print(f"\n📁 OUTPUT FILES:")
        print(f"   1. 04_OA_Altmetrics_merge.xlsx - {self.stats['total_matches']} matched papers")
        print(f"   2. 04_OA_Altmetrics_merge.dta - Stata format")
        print(f"   3. 04_not_matched.xlsx - {self.stats['oa_not_matched'] + self.stats['alt_not_matched']} not matched papers")

        print("\n" + "="*70)
        print("✓ MERGE COMPLETE!")
        print("="*70)


def main():
    """Main execution function"""

    # Define paths
    OPENALEX_PATH = r"C:\Users\User\OneDrive\OneDrive - Universidad de los andes\Global Complexity School\Final project\Excel\03_OpenAlex_API_CS_FULL.xlsx"

    ALTMETRIC_PATH = r"C:\Users\User\OneDrive\OneDrive - Universidad de los andes\Global Complexity School\Final project\Excel\03_Almetrics_API_CS.xlsx"

    OUTPUT_DIR = r"C:\Users\User\OneDrive\OneDrive - Universidad de los andes\Global Complexity School\Final project\Excel"

    # Create merger and run
    merger = DataMerger(OPENALEX_PATH, ALTMETRIC_PATH, OUTPUT_DIR)
    merger.run()


if __name__ == "__main__":
    main()
